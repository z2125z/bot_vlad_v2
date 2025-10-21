import os
from datetime import datetime  
from database.db import Database
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelReportService:
    def __init__(self):
        self.db = Database()
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_comprehensive_report(self):
        """Создать комплексный отчет в Excel без использования pandas"""
        try:
            # Создаем рабочую книгу
            wb = Workbook()
            
            # Удаляем дефолтный лист
            wb.remove(wb.active)
            
            # Создаем листы
            self._create_summary_sheet(wb)
            self._create_users_sheet(wb)
            self._create_activity_sheet(wb)
            self._create_mailings_sheet(wb)
            self._create_analytics_sheet(wb)
            self._create_segments_sheet(wb)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.reports_dir}/bot_report_{timestamp}.xlsx"
            wb.save(filename)
            
            logger.info(f"Excel отчет создан: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Ошибка при создании Excel отчета: {e}")
            return None

    def _create_summary_sheet(self, wb):
        """Лист с общей статистикой"""
        ws = wb.create_sheet("Общая статистика")
        stats = self.db.get_detailed_stats()
        
        # Заголовки
        headers = ['Метрика', 'Значение']
        ws.append(headers)
        
        # Данные
        data = [
            ['Всего пользователей', stats['total_users']],
            ['Новых пользователей (24ч)', stats['new_users_today']],
            ['Новых пользователей (7 дней)', stats['new_users_week']],
            ['Новых пользователей (30 дней)', stats['new_users_month']],
            ['Активных пользователей (24ч)', stats['active_users_today']],
            ['Активных пользователей (7 дней)', stats['active_users_week']],
            ['Всего рассылок', stats['total_mailings']],
            ['Отправлено сообщений', stats['total_sent_messages']],
            ['Успешных доставок', stats['successful_deliveries']],
            ['Неудачных отправок', stats['failed_deliveries']],
            ['Процент доставки', f"{round((stats['successful_deliveries']/(stats['successful_deliveries'] + stats['failed_deliveries']))*100, 2) if (stats['successful_deliveries'] + stats['failed_deliveries']) > 0 else 0}%"],
            ['Средняя активность на пользователя', stats['avg_activity_per_user']]
        ]
        
        for row in data:
            ws.append(row)
        
        # Стилизация
        self._apply_styles(ws, 1)  # Изменено с 2 на 1, так как заголовок в первой строке

    def _create_users_sheet(self, wb):
        """Лист с данными пользователей"""
        ws = wb.create_sheet("Пользователи")
        users = self.db.get_all_users()
        
        # Заголовки
        headers = ['ID', 'Username', 'Имя', 'Фамилия', 'Дата регистрации', 'Последняя активность']
        ws.append(headers)
        
        # Данные
        for user in users:
            ws.append([
                user['user_id'],
                user['username'] or 'Не указан',
                user['first_name'] or 'Не указано',
                user['last_name'] or 'Не указана',
                user['joined_date'],
                user['last_activity']
            ])
        
        self._apply_styles(ws, 1)

    def _create_activity_sheet(self, wb):
        """Лист с активностью пользователей"""
        ws = wb.create_sheet("Активность")
        top_users = self.db.get_top_active_users(50)
        
        # Заголовки
        headers = ['User ID', 'Username', 'Имя', 'Количество действий']
        ws.append(headers)
        
        # Данные
        for user in top_users:
            ws.append([
                user['user_id'],
                user['username'] or 'Не указан',
                user['first_name'] or 'Не указано',
                user['activity_count']
            ])
        
        self._apply_styles(ws, 1)
        
        # Лист с активностью по дням
        ws_daily = wb.create_sheet("Активность по дням")
        activity_data = self.db.get_activity_data(30)
        
        ws_daily.append(['Дата', 'Количество действий'])
        for date, count in activity_data:
            ws_daily.append([date, count])
        
        self._apply_styles(ws_daily, 1)

    def _create_mailings_sheet(self, wb):
        """Лист с данными рассылок"""
        ws = wb.create_sheet("Рассылки")
        mailings = self.db.get_all_mailings()
        
        # Заголовки
        headers = ['ID', 'Заголовок', 'Тип', 'Аудитория', 'Дата создания', 'Отправлено сообщений', 'Доставлено']
        ws.append(headers)
        
        # Данные
        for mailing in mailings:
            ws.append([
                mailing['id'],
                mailing['title'] or 'Без заголовка',
                mailing['message_type'],
                mailing['audience_type'],
                mailing['created_at'],
                mailing['sent_count'],
                mailing['delivered_count']
            ])
        
        self._apply_styles(ws, 1)
        
        # Лист с эффективностью рассылок
        ws_perf = wb.create_sheet("Эффективность рассылок")
        mailing_performance = self.db.get_mailing_performance()
        
        ws_perf.append(['ID рассылки', 'Заголовок', 'Дата', 'Отправлено', 'Доставлено', 'Процент доставки'])
        for perf in mailing_performance:
            ws_perf.append([
                perf['id'],
                perf['title'],
                perf['created_at'],
                perf['sent_count'],
                perf['delivered_count'],
                f"{perf['delivery_rate']}%"
            ])
        
        self._apply_styles(ws_perf, 1)

    def _create_analytics_sheet(self, wb):
        """Лист с аналитикой"""
        ws = wb.create_sheet("Рост пользователей")
        growth_data = self.db.get_user_growth_data(30)
        
        ws.append(['Дата', 'Новых пользователей', 'Общее количество'])
        
        cumulative = 0
        for date, count in growth_data:
            cumulative += count
            ws.append([date, count, cumulative])
        
        self._apply_styles(ws, 1)

    def _create_segments_sheet(self, wb):
        """Лист с сегментами пользователей"""
        ws = wb.create_sheet("Сегменты пользователей")
        segments = self.db.get_user_segments()
        total_users = self.db.get_user_count()
        
        ws.append(['Сегмент', 'Количество', 'Процент от общего числа'])
        
        segments_data = [
            ['Новые пользователи (7 дней)', segments['new_users'], f"{(segments['new_users']/total_users)*100:.2f}%" if total_users > 0 else "0%"],
            ['Активные пользователи (7 дней)', segments['active_users'], f"{(segments['active_users']/total_users)*100:.2f}%" if total_users > 0 else "0%"],
            ['Неактивные пользователей (30+ дней)', segments['inactive_users'], f"{(segments['inactive_users']/total_users)*100:.2f}%" if total_users > 0 else "0%"],
            ['Пользователи с username', segments['users_with_username'], f"{(segments['users_with_username']/total_users)*100:.2f}%" if total_users > 0 else "0%"],
            ['Пользователи без username', segments['users_without_username'], f"{(segments['users_without_username']/total_users)*100:.2f}%" if total_users > 0 else "0%"]
        ]
        
        for row in segments_data:
            ws.append(row)
        
        self._apply_styles(ws, 1)

    def _apply_styles(self, worksheet, header_row):
        """Применить стили к листу"""
        # Стиль для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Применяем стили к заголовкам
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=header_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Автоподбор ширины колонок
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    def cleanup_old_reports(self, keep_last=5):
        """Удалить старые отчеты, оставить только последние keep_last"""
        try:
            reports = []
            for filename in os.listdir(self.reports_dir):
                if filename.startswith("bot_report_") and filename.endswith(".xlsx"):
                    filepath = os.path.join(self.reports_dir, filename)
                    reports.append((filepath, os.path.getctime(filepath)))
            
            # Сортируем по дате создания (новые первыми)
            reports.sort(key=lambda x: x[1], reverse=True)
            
            # Удаляем старые отчеты
            for filepath, _ in reports[keep_last:]:
                os.remove(filepath)
                logger.info(f"Удален старый отчет: {filepath}")
                
        except Exception as e:
            logger.error(f"Ошибка при очистке старых отчетов: {e}")